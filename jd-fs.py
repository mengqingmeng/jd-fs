##coding=utf-8
from urllib import request,parse
from bs4 import BeautifulSoup
import json,logging,os
from openpyxl import Workbook
keyword = "防晒"


fromIndex = input("输入开始页数：")
toIndex = input("输入结束页数：")

def cutStr(str):

    return str

def getData(fromIndex,toIndex):
    exits = os.path.exists('example.xls')
    workbook = Workbook()
    sheet =  workbook.create_sheet("jd-fs", 0)
    sheet.cell(row=1,column=1).value = "名称"
    sheet.cell(row=1,column=2).value = "价格"
    sheet.cell(row =1, column=3).value = "链接"
    sheet.cell(row=1,column=4).value = "评论数"
    line = 2;
    for index in range(fromIndex,toIndex):

        print("页数：",index)
        baseurl = "https://search.jd.com/Search?keyword=" + parse.quote(
            keyword) + "&enc=utf-8&qrst=1&rt=1&stop=1&vt=2&suggest=1.his.0.0&psort=3&wtype=1&click=1&page="+str(2*index-1)+"&s="+str(30*(index-1)+1)
        #print(baseurl)
        req = request.Request(baseurl)
        req.add_header('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                                     'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36')
        try:
            with request.urlopen(req) as f:
                returnData =  f.read().decode('utf-8')
                soup = BeautifulSoup(returnData,"lxml")
                for product in soup.find_all("li", class_="gl-item"):
                    price = product.find("div",class_="p-price").find('i').get_text()
                    name = product.find("div",class_="p-name").find('em').get_text()
                    if "【京东超市】" in name:
                        name = name.replace("【京东超市】"," ")
                    url = product.find("div",class_="p-name").find('a').get('href')
                    comments = product.find("div",class_="p-commit").find('a').get_text()

                    lineDatas = [name,price,url,comments]

                    for l in range(len(lineDatas)):
                        sheet.cell(row=line, column=l+1).value=lineDatas[l]
                    line = line + 1
        except Exception as err:
            print(err)
            logging.debug("fail 页数: %d %s", index, err)

    workbook.save('jd.xlsx')


getData(int(fromIndex),int(toIndex))
#getData(1,2)

